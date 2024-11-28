import pandas as pd
from tkinter import Tk, filedialog, Label, Button, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os


class PlanilhaVendasETaxasApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Processador de Vendas e Taxas v1.0")
        self.root.geometry("500x400")

        # Seção de Upload da Planilha de Vendas
        Label(root, text="Selecione a planilha de vendas:").pack(pady=20)
        self.label_vendas = Label(root, text="Nenhum arquivo selecionado", fg="gray")
        self.label_vendas.pack(pady=10)
        Button(root, text="Upload Planilha de Vendas", command=self.selecionar_planilha_vendas).pack(pady=5)

        # Seção de Upload da Planilha de Taxas
        Label(root, text="Selecione a planilha de taxas:").pack(pady=10)
        self.label_taxas = Label(root, text="Nenhum arquivo selecionado", fg="gray")
        self.label_taxas.pack(pady=5)
        Button(root, text="Upload Planilha de Taxas", command=self.selecionar_planilha_taxas).pack(pady=5)

        # Barra de Progresso
        self.label_status = Label(root, text="", fg="green", wraplength=400, justify="center")
        self.label_status.pack(pady=10)

        self.progress_bar = ttk.Progressbar(root, orient="horizontal", length=400, mode="determinate")
        self.progress_bar.pack(pady=10)

        # Botão para Processar as Planilhas
        Button(root, text="Iniciar Processamento", command=self.processar_planilha).pack(pady=10)

        # Inicializando variáveis dos arquivos
        self.arquivo_vendas = None
        self.arquivo_taxas = None

    def selecionar_planilha_vendas(self):
        self.arquivo_vendas = filedialog.askopenfilename(
            title="Selecione a planilha de vendas", filetypes=[("Excel files", "*.xlsx")]
        )
        if self.arquivo_vendas:
            self.label_vendas.config(text=os.path.basename(self.arquivo_vendas))

    def selecionar_planilha_taxas(self):
        self.arquivo_taxas = filedialog.askopenfilename(
            title="Selecione a planilha de taxas", filetypes=[("Excel files", "*.xlsx")]
        )
        if self.arquivo_taxas:
            self.label_taxas.config(text=os.path.basename(self.arquivo_taxas))

    def processar_planilha(self):
        if not self.arquivo_vendas or not self.arquivo_taxas:
            messagebox.showerror("Erro", "Selecione ambas as planilhas antes de iniciar o processamento.")
            return

        try:
            self.label_status.config(text="Processando...")
            self.progress_bar["value"] = 0
            self.root.update_idletasks()

            # Carregar as planilhas
            df_vendas = pd.read_excel(self.arquivo_vendas)
            df_taxas = pd.read_excel(self.arquivo_taxas)

            # Normalizar datas
            df_vendas["Data"] = pd.to_datetime(df_vendas["Data"], errors="coerce")
            df_taxas["Data Inicial"] = pd.to_datetime(df_taxas["Data Inicial"], errors="coerce")
            df_taxas["Data Final"] = pd.to_datetime(df_taxas["Data Final"], errors="coerce")

            # Consultar e preencher taxas
            df_vendas_com_taxas = self.consulta_e_preenche_taxas(
                df_vendas, df_taxas,
                col_cartoes_vendas="Cartões", col_cartoes_taxas="Cartão",
                col_data_vendas="Data", col_data_inicial="Data Inicial", col_data_final="Data Final",
                col_parcelas_vendas="Total de Parcelas", col_parcelas_taxas="Parcelas",
                col_taxa="Taxa"
            )

            # Remover a coluna "Taxa Correspondente" e mover os dados para a Coluna S com título "Comissão Contratada"
            if "Taxa Correspondente" in df_vendas_com_taxas.columns:
                df_vendas_com_taxas.rename(columns={"Taxa Correspondente": "Comissão Contratada"}, inplace=True)

            self.progress_bar["value"] = 50
            self.root.update_idletasks()

            # Gerar nome incremental para o arquivo de saída
            nome_arquivo = self.salvar_com_nome_incremental("planilha_calculo")

            # Salvando o arquivo processado com formatação
            self.salvar_com_formatacao(df_vendas_com_taxas, nome_arquivo)

            self.progress_bar["value"] = 100
            self.label_status.config(
                text=f"Sucesso! Salvo como:\n'{os.path.abspath(nome_arquivo)}'",
                fg="green"
            )
            messagebox.showinfo("Sucesso", f"Arquivo processado com sucesso! Salvo como '{nome_arquivo}'.")
            self.zerar_estado()
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao processar as planilhas: {e}")
            print(f"Erro ao processar as planilhas: {e}")

    def consulta_e_preenche_taxas(self, df_vendas, df_taxas, col_cartoes_vendas, col_cartoes_taxas, col_data_vendas,
                                  col_data_inicial, col_data_final, col_parcelas_vendas, col_parcelas_taxas, col_taxa):
        """
        Preenche o DataFrame de Vendas com as taxas correspondentes baseando-se nas condições de match.
        """
        df_vendas['Taxa Correspondente'] = None

        for index, row in df_vendas.iterrows():
            taxa_match = df_taxas[
                (df_taxas[col_data_inicial] <= row[col_data_vendas]) &
                (df_taxas[col_data_final] >= row[col_data_vendas]) &
                (df_taxas[col_parcelas_taxas] == row[col_parcelas_vendas]) &
                (df_taxas[col_cartoes_taxas] == row[col_cartoes_vendas])
            ]
            if not taxa_match.empty:
                df_vendas.at[index, 'Taxa Correspondente'] = taxa_match.iloc[0][col_taxa]

        return df_vendas

    def salvar_com_nome_incremental(self, nome_base):
        contador = 1
        nome_arquivo = f"{nome_base}({contador}).xlsx"
        while os.path.exists(nome_arquivo):
            contador += 1
            nome_arquivo = f"{nome_base}({contador}).xlsx"
        return nome_arquivo

    def salvar_com_formatacao(self, dataframe, output_file):
        dataframe.to_excel(output_file, index=False, engine="openpyxl")
        wb = load_workbook(output_file)
        ws = wb.active


     # Adicionar fórmula "=M-O" na coluna correspondente
        col_valor_bruto = "M"  # Supondo que "Valor Bruto" está na coluna M
        col_valor_liquido = "O"  # Supondo que "Valor Líquido" está na coluna O
        col_retenção = "R"  # Próxima coluna para o resultado (ajuste conforme necessário)

        ws[f"{col_retenção}1"] = "Retenção"  # Cabeçalho da nova coluna

        for row in range(2, ws.max_row + 1):  # Começa na linha 2 para evitar o cabeçalho
            ws[f"{col_retenção}{row}"].value = f"={col_valor_bruto}{row}-{col_valor_liquido}{row}"

        # Nova Coluna Comissão Aplicada
        col_valor_bruto = "M"  # Supondo que "Valor Bruto" está na coluna M
        col_retenção = "R"  # Supondo que "Retenção" está na coluna P
        col_comissão_aplicada = "S"  # Nova coluna para comissão aplicada

        # Adicionando o cabeçalho da nova coluna
        ws[f"{col_comissão_aplicada}1"] = "Comissão Aplicada"  # Cabeçalho da nova coluna

        # Preenchendo a fórmula em cada linha da coluna col_comissão_aplicada
        for row in range(2, ws.max_row + 1):  # Começa na linha 2 para evitar o cabeçalho
            ws[f"{col_comissão_aplicada}{row}"].value = f"={col_retenção}{row}*100/{col_valor_bruto}{row}"

        # Nova Coluna Valor Líquido Contratado
        col_taxa = "Q"  # Supondo que "Taxa" está na coluna Q
        col_valor_bruto = "M"  # Supondo que "Valor Bruto" está na coluna M
        col_valor_liquido_contratado = "T"

        # Adicionando o cabeçalho da nova coluna
        ws[f"{col_valor_liquido_contratado}1"] = "Valor Líquido Contatado"  # Cabeçalho da nova coluna

        # Preenchendo a fórmula em cada linha da coluna col_valor_liquido_contratado
        for row in range(2, ws.max_row + 1):  # Começa na linha 2 para evitar o cabeçalho
            ws[f"{col_valor_liquido_contratado}{row}"].value = f"=(100-{col_taxa}{row})/100*{col_valor_bruto}{row}"

        # Nova Coluna Diferença
        col_valor_liquido_contratado = "T"  # Valor do Valor Liquido Contratado
        col_valor_liquido = "O"  # Supondo que "Valor Bruto" está na coluna M
        col_diferença = "U"

        # Adicionando o cabeçalho da nova coluna
        ws[f"{col_diferença}1"] = "Diferença"  # Cabeçalho da nova coluna

        # Preenchendo a fórmula da coluna "Diferença"
        for row in range(2, ws.max_row + 1):  # Começa na linha 2 para evitar o cabeçalho
            ws[f"{col_diferença}{row}"].value = f"={col_valor_liquido_contratado}{row}-{col_valor_liquido}{row}"

        # Nova Coluna Indébito
        col_indébito = "V"

        # Adicionando o cabeçalho da nova coluna
        ws[f"{col_indébito}1"] = "Indébito"  # Cabeçalho da nova coluna

        # Preenchendo a fórmula condicional na coluna "Indébito"
        for row in range(2, ws.max_row + 1):  # Começa na linha 2 para evitar o cabeçalho
            ws[f"{col_indébito}{row}"].value = f"=IF({col_diferença}{row}<=0,0,{col_diferença}{row})"
        # Formatar a altura da linha 1 (cabeçalho)
        ws.row_dimensions[1].height = 48

        # Formatação de cabeçalho
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="7FD4DE", end_color="7FD4DE", fill_type="solid")
            cell.border = thin_border

        # Formatar a Coluna D para exibir datas no formato dd/mm/yyyy
        for row in ws.iter_rows(min_col=4, max_col=4, min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.number_format = "dd/mm/yyyy"

        # Ajustar largura de todas as colunas
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)  # Letra da coluna
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(output_file)

    def zerar_estado(self):
        self.arquivo_vendas = None
        self.arquivo_taxas = None
        self.label_vendas.config(text="Nenhum arquivo selecionado")
        self.label_taxas.config(text="Nenhum arquivo selecionado")
        self.label_status.config(text="")
        self.progress_bar["value"] = 0


root = Tk()
app = PlanilhaVendasETaxasApp(root)
root.mainloop()